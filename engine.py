import numpy as np
from tqdm import tqdm
import torch
from torch.cuda.amp import autocast as autocast
from sklearn.metrics import confusion_matrix, accuracy_score
from utils import save_imgs
from sklearn.metrics import cohen_kappa_score
from sklearn.metrics import mean_absolute_error
import openpyxl
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import worksheet
from configs.config_setting import setting_config

# 后端设置优化
torch.backends.cudnn.deterministic = True  # 保证结果可重复
torch.backends.cudnn.benchmark = True  # 自动选择高效算法，加速
torch.backends.cuda.matmul.allow_tf32 = True  # 启用TF32加速矩阵运算
torch.backends.cuda.matmul.allow_fp16_reduced_precision_reduction = True  # 启用FP16精度优化


def train_one_epoch(train_loader,
                    model,
                    criterion,
                    optimizer,
                    scheduler,
                    epoch,
                    logger,
                    config,
                    scaler=None):

    # 确保使用命令行参数覆盖数据集配置
    if hasattr(config, 'datasets') and config.datasets is not None:
        print(f"Using dataset from config: {config.datasets}")
    else:
        print(f"Using dataset from command line input: {setting_config.datasets}")


    '''
    train model for one epoch
    '''
    # switch to train mode
    model.train()

    loss_list = []

    for iter, data in enumerate(train_loader):
        optimizer.zero_grad()
        images, targets = data
        images, targets = images.cuda(non_blocking=True).float(), targets.cuda(non_blocking=True).float()
        if config.amp:
            with autocast():
                out = model(images)
                loss = criterion(out, targets)
            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()
        else:
            out = model(images)
            loss = criterion(out, targets)
            loss.backward()
            optimizer.step()

        loss_list.append(loss.item())

        now_lr = optimizer.state_dict()['param_groups'][0]['lr']
        if iter % config.print_interval == 0:
            log_info = f'train: epoch {epoch}, iter:{iter}, loss: {np.mean(loss_list):.4f}, lr: {now_lr}'
            print(log_info)
            logger.info(log_info)
    scheduler.step()


def val_one_epoch(test_loader,
                  model,
                  criterion,
                  epoch,
                  logger,
                  config):

    # 确保使用命令行参数覆盖数据集配置
    if hasattr(config, 'datasets') and config.datasets is not None:
        print(f"Using dataset from config: {config.datasets}")
    else:
        print(f"Using dataset from command line input: {setting_config.datasets}")
    # switch to evaluate mode
    model.eval()
    preds = []
    gts = []
    loss_list = []
    with torch.no_grad():
        for data in tqdm(test_loader):
            img, msk = data
            img, msk = img.cuda(non_blocking=True).float(), msk.cuda(non_blocking=True).float()
            out = model(img)
            loss = criterion(out, msk)
            loss_list.append(loss.item())
            gts.append(msk.squeeze(1).cpu().detach().numpy())
            if type(out) is tuple:
                out = out[0]
            out = out.squeeze(1).cpu().detach().numpy()
            preds.append(out)

    if epoch % config.val_interval == 0:
        preds = np.array(preds).reshape(-1)
        gts = np.array(gts).reshape(-1)

        y_pre = np.where(preds >= config.threshold, 1, 0)
        y_true = np.where(gts >= 0.5, 1, 0)

        confusion = confusion_matrix(y_true, y_pre)
        TN, FP, FN, TP = confusion[0, 0], confusion[0, 1], confusion[1, 0], confusion[1, 1]

        accuracy = float(TN + TP) / float(np.sum(confusion)) if float(np.sum(confusion)) != 0 else 0
        sensitivity = float(TP) / float(TP + FN) if float(TP + FN) != 0 else 0
        specificity = float(TN) / float(TN + FP) if float(TN + FP) != 0 else 0
        f1_or_dsc = float(2 * TP) / float(2 * TP + FP + FN) if float(2 * TP + FP + FN) != 0 else 0
        miou = float(TP) / float(TP + FP + FN) if float(TP + FP + FN) != 0 else 0

        # 本人后加的
        iou = TP / (TP + FN + FP + 1e-7)
        precision = TP / (TP + FP + 1e-7)  # 计算precision（精确率）
        recall = TP / (TP + FN + 1e-7)  # 召回率recall 灵敏度sensitivity的计算公式是一样的
        f1 = 2 * (precision * recall) / (precision + recall + 1e-7)  # 计算f1-score
        mae = mean_absolute_error(y_true.flatten(), y_pre.flatten())

        kappa = cohen_kappa_score(y_true, y_pre)  # 计算Kappa系数
        oa = accuracy_score(y_true, y_pre)

        # pa mpa借鉴  https://blog.csdn.net/qq_41375318/article/details/108380694
        '''https://blog.csdn.net/weixin_39919165/article/details/110313831?ops_request_misc=%257B%2522request%255Fid%2522%253A%2522171369138216800211590323%2522%252C%2522scm%2522%253A%252220140713.130102334.pc%255Fall.%2522%257D&request_id=171369138216800211590323&biz_id=0&utm_medium=distribute.pc_search_result.none-task-blog-2~all~first_rank_ecpm_v1~rank_v31_ecpm-2-110313831-null-null.142^v100^pc_search_result_base5&utm_term=%E5%9B%BE%E5%83%8F%E5%88%86%E5%89%B2%E6%B1%82oa%E7%9A%84%E4%BB%A3%E7%A0%81&spm=1018.2226.3001.4187
        '''

        def Pixel_Accuracy(confusion_matrix):
            Acc = np.diag(confusion_matrix).sum() / confusion_matrix.sum()
            return Acc

        pa = Pixel_Accuracy(confusion)  # 求pa

        def Pixel_Accuracy_Class(confusion_matrix):
            Acc = np.diag(confusion_matrix) / confusion_matrix.sum(axis=1)
            Acc = np.nanmean(Acc)
            return Acc

        mpa = Pixel_Accuracy_Class(confusion)  # 求mpa

        #

        log_info = f'val epoch: {epoch}, loss: {np.mean(loss_list):.4f}, miou: {miou}, f1_or_dsc: {f1_or_dsc}, accuracy: {accuracy}, \
                specificity: {specificity}, sensitivity: {sensitivity}, confusion_matrix: {confusion}'
        print(log_info)
        logger.info(log_info)

    else:
        log_info = f'val epoch: {epoch}, loss: {np.mean(loss_list):.4f}'
        print(log_info)
        logger.info(log_info)

    return np.mean(loss_list)


def test_one_epoch(test_loader,
                   model,
                   criterion,
                   logger,
                   config,
                   test_data_name=None):
    # 确保使用命令行参数覆盖数据集配置
    if hasattr(config, 'datasets') and config.datasets is not None:
        print(f"Using dataset from config: {config.datasets}")
    else:
        print(f"Using dataset from command line input: {setting_config.datasets}")
    # switch to evaluate mode
    model.eval()
    preds = []
    gts = []
    loss_list = []
    with torch.no_grad():
        for i, data in enumerate(tqdm(test_loader)):
            img, msk = data
            img, msk = img.cuda(non_blocking=True).float(), msk.cuda(non_blocking=True).float()
            out = model(img)
            loss = criterion(out, msk)
            loss_list.append(loss.item())
            msk = msk.squeeze(1).cpu().detach().numpy()
            gts.append(msk)
            if type(out) is tuple:
                out = out[0]
            out = out.squeeze(1).cpu().detach().numpy()
            preds.append(out)
            # save_imgs(img, msk, out, i, config.work_dir + 'outputs/', config.datasets, config.threshold,
            #           test_data_name=test_data_name)

        preds = np.array(preds).reshape(-1)
        gts = np.array(gts).reshape(-1)

        y_pre = np.where(preds >= config.threshold, 1, 0)
        y_true = np.where(gts >= 0.5, 1, 0)

        confusion = confusion_matrix(y_true, y_pre)
        TN, FP, FN, TP = confusion[0, 0], confusion[0, 1], confusion[1, 0], confusion[1, 1]

        accuracy = float(TN + TP) / float(np.sum(confusion)) if float(np.sum(confusion)) != 0 else 0
        sensitivity = float(TP) / float(TP + FN) if float(TP + FN) != 0 else 0
        specificity = float(TN) / float(TN + FP) if float(TN + FP) != 0 else 0
        f1_or_dsc = float(2 * TP) / float(2 * TP + FP + FN) if float(2 * TP + FP + FN) != 0 else 0
        miou = float(TP) / float(TP + FP + FN) if float(TP + FP + FN) != 0 else 0

        # 本人后加的
        iou = TP / (TP + FN + FP + 1e-7)
        precision = TP / (TP + FP + 1e-7)  # 计算precision（精确率）
        recall = TP / (TP + FN + 1e-7)  # 召回率recall 灵敏度sensitivity的计算公式是一样的
        f1 = 2 * (precision * recall) / (precision + recall + 1e-7)  # 计算f1-score
        mae = mean_absolute_error(y_true.flatten(), y_pre.flatten())

        kappa = cohen_kappa_score(y_true, y_pre)  # 计算Kappa系数
        oa = accuracy_score(y_true, y_pre)

        # pa mpa借鉴  https://blog.csdn.net/qq_41375318/article/details/108380694
        '''https://blog.csdn.net/weixin_39919165/article/details/110313831?ops_request_misc=%257B%2522request%255Fid%2522%253A%2522171369138216800211590323%2522%252C%2522scm%2522%253A%252220140713.130102334.pc%255Fall.%2522%257D&request_id=171369138216800211590323&biz_id=0&utm_medium=distribute.pc_search_result.none-task-blog-2~all~first_rank_ecpm_v1~rank_v31_ecpm-2-110313831-null-null.142^v100^pc_search_result_base5&utm_term=%E5%9B%BE%E5%83%8F%E5%88%86%E5%89%B2%E6%B1%82oa%E7%9A%84%E4%BB%A3%E7%A0%81&spm=1018.2226.3001.4187
        '''

        def Pixel_Accuracy(confusion_matrix):
            Acc = np.diag(confusion_matrix).sum() / confusion_matrix.sum()
            return Acc

        pa = Pixel_Accuracy(confusion)  # 求pa

        def Pixel_Accuracy_Class(confusion_matrix):
            Acc = np.diag(confusion_matrix) / confusion_matrix.sum(axis=1)
            Acc = np.nanmean(Acc)
            return Acc

        mpa = Pixel_Accuracy_Class(confusion)  # 求mpa
        #
        if test_data_name is not None:
            log_info = f'test_datasets_name: {test_data_name}'
            print(log_info)
            logger.info(log_info)
        log_info = f'test of best model, loss: {np.mean(loss_list):.4f},miou: {miou}, f1_or_dsc: {f1_or_dsc}, accuracy: {accuracy}, \
                specificity: {specificity}, sensitivity: {sensitivity}, confusion_matrix: {confusion},iou:{iou},precision:{precision},f1-score:{f1},\
                        mae:{mae},kappa:{kappa},pa:{pa},mpa:{mpa},oa:{oa}'

        print(log_info)
        logger.info(log_info)
        # -----------------------------------------------------------------------------------------
        # file_full_path = r'data/result.xlsx'  # 表格自己创建   data\result.xlsx   ../data/result.xlsx
        file_full_path = r'data/result(batch_size=8).xlsx'  # 表格自己创建   data\result.xlsx   ../data/result.xlsx

        if os.path.isfile(file_full_path) != True:
            workbook = Workbook()
            worksheet = workbook.active

            # ws1 = workbook.create_sheet("isic2017", 1)
            worksheet.title = setting_config.datasets
            column_names = ['model', 'iou', 'miou', 'f1_or_dsc', 'accuracy', 'mpa', 'specificity', 'sensitivity',
                            'precision', 'f1', 'mae', 'kappa']
            for i in range(len(column_names)):
                worksheet.cell(row=1, column=i + 1, value=column_names[i])

            workbook.save(file_full_path)

        # sheet名称
        # sheet_name = setting_config.datasets
        sheet_name = config.datasets

        # 获取指定的文件
        wb = openpyxl.load_workbook(file_full_path)
        # 获取指定的sheet
        ws = wb[sheet_name]
        # 获得最大行数
        max_row_num = ws.max_row
        # 获得最大列数
        max_col_num = ws.max_column

        # 将当前行设置为最大行数
        ws._current_row = max_row_num

        # 使用append方法，将行数据按行追加写入
        # model = setting_config.network
        model = config.network+"_"+str(config.batch_size)

        values = [model, iou, miou, f1_or_dsc, accuracy, mpa, specificity, sensitivity, precision, f1, mae, kappa]

        ws.append(values)

        # 保存文件
        wb.save(file_full_path)
        print('{}  写入excel表格成功'.format(model))

        # -----------------------------------------------------------------------------------------------------------


        if test_data_name is not None:
            log_info = f'test_datasets_name: {test_data_name}'
            print(log_info)
            logger.info(log_info)
        log_info = f'test of best model, loss: {np.mean(loss_list):.4f},miou: {miou}, f1_or_dsc: {f1_or_dsc}, accuracy: {accuracy}, \
                specificity: {specificity}, sensitivity: {sensitivity}, confusion_matrix: {confusion}'
        print(log_info)
        logger.info(log_info)

    return np.mean(loss_list)